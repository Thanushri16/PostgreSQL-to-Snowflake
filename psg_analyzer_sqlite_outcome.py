''' *******************************************************************************************************

* Product: Snowflake Migration Platform

* Utility: "psg_analyzer_sqlite_outcome" which saves the outcome of the psg_analyzer as tables in SQLite 

* Date: Jun 2021

* Company: Dattendriya Data Science Solutions

*********************************************************************************************************** '''

import sqlite3
import os
import pandas as pd
from openpyxl import load_workbook

path=os.getcwd()
db_conn = sqlite3.connect(path+'\\psg_analyzer_outcome.db')
c = db_conn.cursor()

wb = load_workbook(path+'\\psg_meta_info.xlsx', read_only=True, keep_links=False)
all_sheets=wb.sheetnames
print(all_sheets)
all_sheets=all_sheets[:-1]
print(all_sheets)

for sheet in all_sheets:
    df=pd.read_excel(path+'\\psg_meta_info.xlsx', sheet_name=sheet)
    df.to_sql(sheet, db_conn, if_exists='append', index=False)
    print('\n\n***************'+sheet+'***************')
    print(pd.read_sql("SELECT * FROM "+sheet+" LIMIT 5;", db_conn)) 

    

